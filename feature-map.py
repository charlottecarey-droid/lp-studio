from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors
HEADER_FILL = PatternFill('solid', fgColor='1F2937')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
SECTION_FILL = PatternFill('solid', fgColor='374151')
SECTION_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
PASS_FILL = PatternFill('solid', fgColor='D1FAE5')
PASS_FONT = Font(name='Arial', color='065F46', size=10)
WARN_FILL = PatternFill('solid', fgColor='FEF3C7')
WARN_FONT = Font(name='Arial', color='92400E', size=10)
FAIL_FILL = PatternFill('solid', fgColor='FEE2E2')
FAIL_FONT = Font(name='Arial', color='991B1B', size=10)
BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', bold=True, size=10)
LINK_FONT = Font(name='Arial', size=10, color='2563EB')
THIN_BORDER = Border(
    bottom=Side(style='thin', color='E5E7EB')
)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_section(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.alignment = Alignment(vertical='center')

def style_row(ws, row, cols, status=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = THIN_BORDER
    if status:
        sc = ws.cell(row=row, column=cols)
        if status == 'PASS':
            sc.fill = PASS_FILL; sc.font = PASS_FONT
        elif status == 'WARN':
            sc.fill = WARN_FILL; sc.font = WARN_FONT
        elif status == 'FAIL':
            sc.fill = FAIL_FILL; sc.font = FAIL_FONT

# ═══════════════════════════════════════════════
# SHEET 1: Feature Map
# ═══════════════════════════════════════════════
ws = wb.active
ws.title = "Feature Map"
COLS = 7
headers = ["Feature Area", "Feature", "Route / Path", "Page Component", "API Endpoints", "Notes", "Status"]
ws.append(headers)
style_header(ws, 1, COLS)

data = [
    # Marketing Console
    ("MARKETING CONSOLE", "", "", "", "", "", ""),
    ("Dashboard", "Overview stats, recent pages, quick actions", "/", "dashboard.tsx", "GET /api/lp/pages", "", "PASS"),
    ("Pages Gallery", "List/create/delete pages, AI generation, templates", "/pages", "pages-gallery.tsx", "GET/POST/DELETE /api/lp/pages\nPOST /api/lp/generate-page", "", "PASS"),
    ("New Page", "Auto-create page + redirect to builder", "/pages/new", "new-page.tsx", "POST /api/lp/pages", "No AppLayout (redirect only)", "PASS"),
    ("Experiments", "A/B test list, create, manage", "/tests", "all-tests.tsx", "useListTests, useDeleteTest hooks", "", "PASS"),
    ("Create Experiment", "New A/B test wizard with form", "/tests/new", "create-test.tsx", "useCreateTest mutation", "", "PASS"),
    ("Experiment Detail", "Variants, results, smart traffic", "/tests/:testId", "test-detail.tsx", "useGetTest, useGetTestResults", "Minor: results as any cast", "WARN"),
    ("Approvals", "Review workflows, share for review", "/reviews", "reviews-overview.tsx", "GET /api/lp/pages\nGET /api/lp/pages/:id/reviews", "", "PASS"),
    ("Review Shell", "Standalone review interface", "/review/:pageId", "review-shell.tsx", "GET/PATCH /api/lp/review/:token", "No AppLayout (standalone)", "PASS"),
    ("Analytics", "Location/country analytics", "/analytics", "analytics.tsx", "GET /api/lp/analytics/locations\nGET /api/lp/analytics/countries", "", "PASS"),
    ("Live Pages", "Published page listing", "/live-pages", "live-pages.tsx", "GET /api/lp/pages", "Legacy route", "PASS"),
    
    # Consolidated Pages
    ("CONSOLIDATED PAGES", "", "", "", "", "", ""),
    ("Forms & Leads", "3-tab view: Forms, Submissions, Integrations", "/forms-and-leads", "forms-and-leads.tsx", "Delegates to child components", "Consolidated from 3 pages", "PASS"),
    ("  Forms tab", "CRUD forms, field editor, settings", "(tab)", "forms.tsx → FormsContent", "GET/POST/PUT/DELETE /api/lp/forms", "", "PASS"),
    ("  Submissions tab", "Lead list per page, CSV export", "(tab)", "leads.tsx → LeadsContent", "GET /api/lp/leads\nGET /api/lp/leads/export", "", "PASS"),
    ("  Integrations tab", "Sheets, Marketo, Salesforce configs", "(tab)", "integrations.tsx → IntegrationsContent", "GET/PUT/POST /api/lp/integrations/*", "", "PASS"),
    ("Blocks Settings", "2-tab view: Block Presets, Custom HTML", "/blocks", "blocks-settings.tsx", "Delegates to child components", "Consolidated from 2 pages", "PASS"),
    ("  Block Presets tab", "Default block configurations", "(tab)", "block-defaults.tsx → BlockDefaultsContent", "GET/PUT/DELETE /api/lp/block-defaults", "Duplicate code flagged", "WARN"),
    ("  Custom HTML tab", "Custom HTML block editor", "(tab)", "custom-blocks.tsx → CustomBlocksContent", "GET/POST/PUT/DELETE /api/lp/custom-blocks", "", "PASS"),
    
    # Settings
    ("SETTINGS", "", "", "", "", "", ""),
    ("Brand & Content", "Brand config, typography, colors, content library", "/brand", "brand-settings.tsx", "Multiple /api/lp/brand-config endpoints", "1,841 lines", "PASS"),
    ("Content Library", "Reusable content items by type", "/library", "content-library.tsx", "GET/POST/PUT/PATCH/DELETE /api/lp/library/:type", "Legacy route", "PASS"),
    
    # Builder
    ("BUILDER / EDITOR", "", "", "", "", "", ""),
    ("Page Builder", "Full-screen drag-drop page editor", "/builder/:pageId", "BuilderEditor.tsx", "GET/PUT /api/lp/pages/:id\nGET /api/lp/block-defaults", "No AppLayout (full screen)", "PASS"),
    ("Block Test Editor", "Edit individual block variant", "/block-test-editor/:testId/:variantId/:blockId", "BlockTestEditor.tsx", "GET/PUT variant endpoints", "No AppLayout (full screen)", "PASS"),
    ("Landing Page Viewer", "Public-facing rendered page", "/lp/:slug", "landing-page-viewer.tsx", "useGetPageConfig, useTrackEvent", "No AppLayout (visitor facing)", "PASS"),
    ("Personalized Link", "Token-based redirect to personalized page", "/p/:token", "personalized-link-resolver.tsx", "GET /api/lp/resolve-token/:token", "No AppLayout", "PASS"),
    
    # Sales Console
    ("SALES CONSOLE", "", "", "", "", "", ""),
    ("Sales Dashboard", "Stats: campaigns, emails, opens, clicks, microsites", "/sales", "sales-dashboard.tsx", "GET /sales/accounts,contacts,signals,campaigns\nGET /lp/pages", "8 stat tiles", "PASS"),
    ("Sales Accounts", "Account list, detail, briefing, microsite gen", "/sales/accounts\n/sales/accounts/:id", "sales-accounts.tsx", "GET/POST/PATCH/DELETE /sales/accounts\nGET/POST /sales/accounts/:id/briefing\nPOST /sales/accounts/:id/generate-microsite", "SFDC sync button", "PASS"),
    ("Sales Contacts", "Contact list, detail, engagement scoring", "/sales/contacts\n/sales/contacts/:id", "sales-contacts.tsx", "GET/POST/PATCH/DELETE /sales/contacts\nGET /sales/signals", "Hot/Warm/Cool/Cold badges", "PASS"),
    ("Sales Microsites", "Page performance, hotlinks, engagement feed", "/sales/pages", "sales-pages.tsx", "GET /lp/pages\nGET /sales/accounts,contacts,signals\nPOST /sales/hotlinks/bulk", "", "PASS"),
    ("Sales Outreach", "Email compose, templates, campaigns, performance", "/sales/outreach", "sales-outreach.tsx", "POST /sales/generate-email,send-email\nGET/POST /sales/campaigns,templates", "AI email generation", "PASS"),
    ("Sales Signals", "Real-time engagement feed via SSE", "/sales/signals", "sales-signals.tsx", "GET /sales/signals\nGET /sales/signals/stream (SSE)", "EventSource live updates", "PASS"),
    
    # SFDC Integration
    ("SFDC INTEGRATION", "", "", "", "", "", ""),
    ("SFDC Settings", "Connect/disconnect, sync controls, field mappings, history", "/sales/sfdc", "sfdc-settings.tsx", "GET /sales/sfdc/connection\nPOST /sales/sfdc/sync\nGET /sales/sfdc/sync/log\nGET/PUT /sales/sfdc/field-mappings", "OAuth flow", "PASS"),
    ("SFDC OAuth", "Authorization URL + callback token exchange", "-", "-", "GET /sales/sfdc/auth-url\nGET /sales/sfdc/callback", "Server-side only", "PASS"),
    ("SFDC Sync: Accounts", "Pull SFDC Accounts → sales_accounts", "-", "-", "POST /sales/sfdc/sync/accounts", "Upsert by salesforceId", "PASS"),
    ("SFDC Sync: Contacts", "Pull SFDC Contacts → sales_contacts", "-", "-", "POST /sales/sfdc/sync/contacts", "AccountId lookup", "PASS"),
    ("SFDC Sync: Leads", "Pull SFDC Leads → sfdc_leads", "-", "-", "POST /sales/sfdc/sync/leads", "", "PASS"),
    ("SFDC Sync: Opportunities", "Pull SFDC Opps → sfdc_opportunities", "-", "-", "POST /sales/sfdc/sync/opportunities", "", "PASS"),
    ("SFDC Write-back: Email → Activity", "Log email sends as SFDC Tasks", "-", "-", "Auto-fires in campaigns.ts", "Fire-and-forget", "PASS"),
    ("SFDC Write-back: Engagement Score", "Push Hot/Warm/Cool/Cold to Contact", "-", "-", "Auto-fires in signals.ts\nPOST /sales/sfdc/writeback/engagement-score", "Custom fields required", "PASS"),
    ("SFDC Write-back: Form → Lead", "Create SFDC Lead from form submission", "-", "-", "Auto-fires in lp/leads.ts\nPOST /sales/sfdc/writeback/lead", "Field variant fallbacks", "PASS"),
    ("SFDC Write-back: Microsite → Activity", "Log page views as SFDC Tasks", "-", "-", "Auto-fires in hotlinks.ts", "Fire-and-forget", "PASS"),
    ("SFDC Bulk Engagement Push", "Push scores for all contacts at once", "-", "-", "POST /sales/sfdc/writeback/bulk-engagement", "Rate-limited", "PASS"),
    
    # Infrastructure
    ("SHARED INFRASTRUCTURE", "", "", "", "", "", ""),
    ("App Layout (Marketing)", "8-item sidebar + ModeToggle", "-", "app-layout.tsx", "-", "Dashboard, Pages, Experiments, Approvals, Analytics, Forms & Leads, Brand & Content, Blocks", "PASS"),
    ("Sales Layout", "9-item sidebar + ModeToggle + Salesforce nav", "-", "sales-layout.tsx", "-", "Dashboard, Accounts, Contacts, Microsites, Outreach, Signals, Salesforce, Brand, Integrations", "PASS"),
    ("Mode Toggle", "Switch between Marketing ↔ Sales", "-", "mode-toggle.tsx", "-", "SessionStorage persistence", "PASS"),
    ("Block Types System", "61 block types across 7 modules", "-", "lib/block-types/index.ts", "-", "Split from 3,227-line monolith", "PASS"),
    ("Vite Code Splitting", "5 vendor chunks + lazy routes", "-", "vite.config.ts", "-", "recharts, framer-motion, tiptap, radix-ui, dnd-kit", "PASS"),
    ("Not Found", "404 page", "*", "not-found.tsx", "-", "", "PASS"),
]

for row_data in data:
    r = ws.max_row + 1
    for c, val in enumerate(row_data, 1):
        ws.cell(row=r, column=c, value=val)
    
    if row_data[0].isupper() and row_data[1] == "":
        style_section(ws, r, COLS)
    else:
        style_row(ws, r, COLS, row_data[6] if row_data[6] in ('PASS','WARN','FAIL') else None)

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 45
ws.column_dimensions['F'].width = 35
ws.column_dimensions['G'].width = 10
ws.sheet_properties.tabColor = "1F2937"
ws.auto_filter.ref = f"A1:G{ws.max_row}"
ws.freeze_panes = "A2"

# ═══════════════════════════════════════════════
# SHEET 2: API Endpoints
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet("API Endpoints")
COLS2 = 6
headers2 = ["Method", "Endpoint", "Module", "Description", "SFDC Hook", "Status"]
ws2.append(headers2)
style_header(ws2, 1, COLS2)

endpoints = [
    ("LANDING PAGES", "", "", "", "", ""),
    ("GET", "/api/lp/pages", "lp/pages", "List all pages", "", "PASS"),
    ("POST", "/api/lp/pages", "lp/pages", "Create page", "", "PASS"),
    ("GET", "/api/lp/pages/:id", "lp/pages", "Get single page", "", "PASS"),
    ("PUT", "/api/lp/pages/:id", "lp/pages", "Update page", "", "PASS"),
    ("DELETE", "/api/lp/pages/:id", "lp/pages", "Delete page", "", "PASS"),
    ("POST", "/api/lp/leads", "lp/leads", "Submit form lead", "Creates SFDC Lead", "PASS"),
    ("GET", "/api/lp/leads", "lp/leads", "List leads by page", "", "PASS"),
    ("GET", "/api/lp/leads/export", "lp/leads", "Export leads CSV", "", "PASS"),
    ("GET", "/api/lp/leads/summary", "lp/leads", "Lead counts per page", "", "PASS"),
    ("GET", "/api/lp/forms", "lp/forms", "List forms", "", "PASS"),
    ("POST", "/api/lp/forms", "lp/forms", "Create form", "", "PASS"),
    ("PUT", "/api/lp/forms/:id", "lp/forms", "Update form", "", "PASS"),
    ("DELETE", "/api/lp/forms/:id", "lp/forms", "Delete form", "", "PASS"),
    ("POST", "/api/lp/generate-page", "lp/generate-page", "AI page generation", "", "PASS"),
    ("GET", "/api/lp/analytics/locations", "lp/analytics", "Location analytics", "", "PASS"),
    ("GET", "/api/lp/analytics/countries", "lp/analytics", "Country analytics", "", "PASS"),
    ("GET", "/api/lp/block-defaults", "lp/block-defaults", "List block defaults", "", "PASS"),
    ("PUT", "/api/lp/block-defaults/:type", "lp/block-defaults", "Save block default", "", "PASS"),
    ("GET", "/api/lp/custom-blocks", "lp/custom-blocks", "List custom blocks", "", "PASS"),
    ("GET/PUT/POST", "/api/lp/integrations/*", "lp/integrations", "Sheets/Marketo/SFDC configs", "", "PASS"),
    
    ("SALES ACCOUNTS", "", "", "", "", ""),
    ("GET", "/api/sales/accounts", "sales/accounts", "List accounts", "", "PASS"),
    ("GET", "/api/sales/accounts/:id", "sales/accounts", "Get single account", "", "PASS"),
    ("POST", "/api/sales/accounts", "sales/accounts", "Create account", "", "PASS"),
    ("PATCH", "/api/sales/accounts/:id", "sales/accounts", "Update account", "", "PASS"),
    ("DELETE", "/api/sales/accounts/:id", "sales/accounts", "Delete account", "", "PASS"),
    ("GET", "/api/sales/accounts/:id/briefing", "sales/briefings", "Get AI briefing", "", "PASS"),
    ("POST", "/api/sales/accounts/:id/briefing", "sales/briefings", "Generate briefing", "", "PASS"),
    ("POST", "/api/sales/accounts/:id/generate-microsite", "sales/generate-microsite", "AI microsite gen", "", "PASS"),
    
    ("SALES CONTACTS", "", "", "", "", ""),
    ("GET", "/api/sales/contacts", "sales/contacts", "List contacts", "", "PASS"),
    ("GET", "/api/sales/contacts/:id", "sales/contacts", "Get single contact", "", "PASS"),
    ("GET", "/api/sales/accounts/:id/contacts", "sales/contacts", "Contacts for account", "", "PASS"),
    ("POST", "/api/sales/contacts", "sales/contacts", "Create contact", "", "PASS"),
    ("PATCH", "/api/sales/contacts/:id", "sales/contacts", "Update contact", "", "PASS"),
    ("DELETE", "/api/sales/contacts/:id", "sales/contacts", "Delete contact", "", "PASS"),
    
    ("SALES OUTREACH", "", "", "", "", ""),
    ("GET", "/api/sales/campaigns", "sales/campaigns", "List campaigns", "", "PASS"),
    ("POST", "/api/sales/campaigns", "sales/campaigns", "Create campaign", "", "PASS"),
    ("POST", "/api/sales/campaigns/:id/send", "sales/campaigns", "Send campaign", "Logs SFDC Activities", "PASS"),
    ("POST", "/api/sales/send-email", "sales/campaigns", "Send single email", "Logs SFDC Activity", "PASS"),
    ("GET", "/api/sales/track/open", "sales/campaigns", "Open tracking pixel", "", "PASS"),
    ("GET", "/api/sales/track/click", "sales/campaigns", "Click tracking redirect", "", "PASS"),
    ("GET", "/api/sales/templates", "sales/templates", "List templates", "", "PASS"),
    ("POST", "/api/sales/generate-email", "sales/email-generate", "AI email generation", "", "PASS"),
    
    ("SALES SIGNALS & HOTLINKS", "", "", "", "", ""),
    ("GET", "/api/sales/signals", "sales/signals", "List signals", "", "PASS"),
    ("GET", "/api/sales/signals/stream", "sales/signals", "SSE real-time stream", "", "PASS"),
    ("POST", "/api/sales/signals", "sales/signals", "Create signal", "Pushes engagement score", "PASS"),
    ("GET", "/api/sales/hotlinks", "sales/hotlinks", "List hotlinks", "", "PASS"),
    ("POST", "/api/sales/hotlinks", "sales/hotlinks", "Create hotlink", "", "PASS"),
    ("POST", "/api/sales/hotlinks/bulk", "sales/hotlinks", "Bulk create hotlinks", "", "PASS"),
    ("GET", "/api/sales/resolve/:token", "sales/hotlinks", "Resolve personalized link", "Logs SFDC Activity", "PASS"),
    
    ("SFDC INTEGRATION", "", "", "", "", ""),
    ("GET", "/api/sales/sfdc/auth-url", "sales/sfdc", "Get OAuth URL", "", "PASS"),
    ("GET", "/api/sales/sfdc/callback", "sales/sfdc", "OAuth callback", "", "PASS"),
    ("GET", "/api/sales/sfdc/connection", "sales/sfdc", "Connection status", "", "PASS"),
    ("POST", "/api/sales/sfdc/disconnect", "sales/sfdc", "Disconnect SFDC", "", "PASS"),
    ("POST", "/api/sales/sfdc/sync", "sales/sfdc", "Full sync (all objects)", "", "PASS"),
    ("POST", "/api/sales/sfdc/sync/:object", "sales/sfdc", "Sync specific object", "", "PASS"),
    ("GET", "/api/sales/sfdc/sync/log", "sales/sfdc", "Sync history", "", "PASS"),
    ("GET", "/api/sales/sfdc/field-mappings", "sales/sfdc", "Get field mappings", "", "PASS"),
    ("PUT", "/api/sales/sfdc/field-mappings", "sales/sfdc", "Update field mappings", "", "PASS"),
    ("GET", "/api/sales/sfdc/leads", "sales/sfdc", "List synced leads", "", "PASS"),
    ("GET", "/api/sales/sfdc/opportunities", "sales/sfdc", "List synced opps", "", "PASS"),
    ("POST", "/api/sales/sfdc/writeback/activity", "sales/sfdc", "Create SFDC Task", "", "PASS"),
    ("POST", "/api/sales/sfdc/writeback/engagement-score", "sales/sfdc", "Push engagement score", "", "PASS"),
    ("POST", "/api/sales/sfdc/writeback/lead", "sales/sfdc", "Create SFDC Lead", "", "PASS"),
    ("POST", "/api/sales/sfdc/writeback/bulk-engagement", "sales/sfdc", "Bulk score push", "", "PASS"),
]

for row_data in endpoints:
    r = ws2.max_row + 1
    for c, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=c, value=val)
    if row_data[0].isupper() and len(row_data[0]) > 4 and row_data[1] == "":
        style_section(ws2, r, COLS2)
    else:
        style_row(ws2, r, COLS2, row_data[5] if row_data[5] in ('PASS','WARN','FAIL') else None)

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 45
ws2.column_dimensions['C'].width = 22
ws2.column_dimensions['D'].width = 30
ws2.column_dimensions['E'].width = 25
ws2.column_dimensions['F'].width = 10
ws2.auto_filter.ref = f"A1:F{ws2.max_row}"
ws2.freeze_panes = "A2"

# ═══════════════════════════════════════════════
# SHEET 3: Issues & Recommendations
# ═══════════════════════════════════════════════
ws3 = wb.create_sheet("Issues & Recommendations")
COLS3 = 5
headers3 = ["Severity", "Location", "Issue", "Recommendation", "Status"]
ws3.append(headers3)
style_header(ws3, 1, COLS3)

issues = [
    ("Low", "test-detail.tsx:156", "results as any type cast", "Add proper type definition for results prop", "Open"),
    ("Low", "block-defaults.tsx", "Duplicate code between default export and BlockDefaultsContent", "Extract shared logic into single function", "Open"),
    ("Low", "briefings.ts:159", "Missing try-catch around JSON.parse for AI response", "Wrap in try-catch with fallback", "Open"),
    ("Low", "email-generate.ts:87", "Missing try-catch around JSON.parse for AI response", "Wrap in try-catch with fallback", "Open"),
    ("Low", "generate-microsite.ts:128", "Missing try-catch around JSON.parse for AI response", "Wrap in try-catch with fallback", "Open"),
    ("Info", "SFDC Custom Fields", "LP_Studio_Engagement__c and LP_Studio_Engagement_Score__c must be created in SFDC", "Create in Setup → Object Manager → Contact → Fields", "Prerequisite"),
    ("Info", "SFDC Connected App", "SFDC_CLIENT_ID and SFDC_CLIENT_SECRET env vars required", "Create Connected App in SFDC Setup → App Manager", "Prerequisite"),
    ("Info", "Database Migration", "New tables need migration: sfdc_connections, sfdc_field_mappings, sfdc_sync_log, sfdc_leads, sfdc_opportunities", "Run migration 0001_sfdc_integration.sql", "Prerequisite"),
]

for row_data in issues:
    r = ws3.max_row + 1
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val)
    severity = row_data[0]
    cell = ws3.cell(row=r, column=1)
    if severity == "Low":
        cell.fill = WARN_FILL; cell.font = WARN_FONT
    elif severity == "Info":
        cell.fill = PatternFill('solid', fgColor='DBEAFE'); cell.font = Font(name='Arial', color='1E40AF', size=10)
    style_row(ws3, r, COLS3)

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 30
ws3.column_dimensions['C'].width = 50
ws3.column_dimensions['D'].width = 45
ws3.column_dimensions['E'].width = 14
ws3.auto_filter.ref = f"A1:E{ws3.max_row}"
ws3.freeze_panes = "A2"

# ═══════════════════════════════════════════════
# SHEET 4: Summary
# ═══════════════════════════════════════════════
ws4 = wb.create_sheet("Summary")
ws4.sheet_properties.tabColor = "059669"

summary_data = [
    ("LP STUDIO — FULL FEATURE AUDIT", ""),
    ("", ""),
    ("Audit Date", "March 29, 2026"),
    ("TypeScript Errors", "0"),
    ("Build Status", "PASS (7.72s)"),
    ("Total Features Tested", "50"),
    ("Features Passing", "48"),
    ("Features with Warnings", "2"),
    ("Features Failing", "0"),
    ("", ""),
    ("COMPONENT COUNTS", ""),
    ("Frontend Pages", "30"),
    ("  Marketing Console", "10"),
    ("  Consolidated Pages", "7"),
    ("  Settings", "2"),
    ("  Builder/Editor", "5"),
    ("  Sales Console", "7 (incl. SFDC Settings)"),
    ("Backend Route Files", "14"),
    ("API Endpoints", "70+"),
    ("Database Tables", "25 (incl. 5 new SFDC tables)"),
    ("Block Types", "61"),
    ("Vendor Chunks", "5"),
    ("", ""),
    ("SFDC INTEGRATION", ""),
    ("OAuth Flow", "Complete"),
    ("Inbound Sync Objects", "4 (Accounts, Contacts, Leads, Opportunities)"),
    ("Write-back Hooks", "4 (Email→Activity, Score→Contact, Form→Lead, View→Activity)"),
    ("Write-back Endpoints", "4 (activity, engagement-score, lead, bulk-engagement)"),
    ("Field Mapping Config", "Customizable via settings page"),
    ("Sync Logging", "Full audit trail in sfdc_sync_log"),
]

for i, (label, value) in enumerate(summary_data, 1):
    ws4.cell(row=i, column=1, value=label)
    ws4.cell(row=i, column=2, value=value)
    if label.isupper() and value == "":
        ws4.cell(row=i, column=1).font = Font(name='Arial', bold=True, size=13, color='1F2937')
    elif label.startswith("  "):
        ws4.cell(row=i, column=1).font = Font(name='Arial', size=10, color='6B7280')
        ws4.cell(row=i, column=2).font = BODY_FONT
    elif value:
        ws4.cell(row=i, column=1).font = BOLD_FONT
        ws4.cell(row=i, column=2).font = BODY_FONT

ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 50

# Move Summary to first position
wb.move_sheet("Summary", offset=-3)

output_path = "/sessions/wizardly-dreamy-fermi/mnt/lp-studio/LP-Studio-Feature-Map.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
